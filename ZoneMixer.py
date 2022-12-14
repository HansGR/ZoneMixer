# ZoneMixer: a room randomizer for Final Fantasy 6
from random import randrange
from openpyxl import load_workbook

# zoneEater: code to read in door list information and return room information
def zoneEater(fn, whichRooms = []):
    # Assess the type of file
    flag = [(fn[-4::] == '.txt'), (fn.find('.xls') > -1)]  # logical comparison: [is_txt, is_xls]

    if flag[0]:
        # Read in the door definitions from a .txt file; generate rooms & room counts
        textfile = open(fn, "r")
        array = textfile.readlines()
        while array[0].strip(' ')[0]=='#':
            # Remove comment lines at beginning
            array.pop(0)
        data = [j.split('#')[0].strip(' ').split('_') for j in array]
        desc = [j.split('#')[1] for j in array]

        # Check to see if there are an odd number of doors:  if so, one is an exit from this zone, and we exclude it.
        # By convention, the first door is the zone entrance
        if len(data)%2==1:
            data.pop(0)

        # Create the list of available doors
        doors = [j[0] for j in data]

        # Create the lookup for door -> room
        door_rooms = dict([[j[0], j[1]] for j in data])

        # Create the lookup for door -> description
        door_descr = dict([[data[j][0], desc[j]] for j in range(len(data))])

        # Create the lookup for forced connections
        forcing = {}
        for j in range(len(data)):
            if len(data[j][3]) > 0:
                # Forcing[nob] = [list of valid forced nibs (usually length=1)]
                forcing[data[j][0]] = data[j][3:]

        # Create dictionary for which doors are in each room
        room_ids = list(set([j[1] for j in data]))
        room_doors = {}
        room_counts = {}
        type_ind = {'d': 0,   # Normal doors
                    'o': 1,   # "outs" or trapdoor exits (nobs)
                    'i': 2}   # "ins" or trapdoor entrances (nibs)
        for R in room_ids:
            thisdoors = [[], [], []]
            thiscounts = [0, 0, 0]
            for d in data:
                if d[1] == R:
                    thisdoors[type_ind[d[2]]].append(d[0])
                    thiscounts[type_ind[d[2]]] += 1
            room_doors[R] = thisdoors
            room_counts[R] = thiscounts

    elif flag[1]:
        # Read in the door definitions from a .xls file; generate rooms & room counts
        wb = load_workbook(fn)
        ws = wb['Exits']

        doors = []
        door_rooms = {}
        door_descr = {}
        rooms = []
        room_doors = {}
        room_counts = {}
        forcing = {}

        for r in ws.iter_rows(3):
            this_ID = r[0].value
            # this_parentMap = r[1].value
            # this_X = r[2].value
            # this_Y = r[3].value
            # this_destMap = r[4].value
            # this_destX = r[5].value
            # this_destY = r[6].value
            # this_facing = r[7].value
            # this_displayLocName = r[8].value
            # this_refreshParentMap = r[9].value
            # this_unknown = r[10].value
            # this_longExitSize = r[11].value
            # this_longExitOrientation = r[12].value
            this_name = r[16].value
            this_roomID = r[22].value
            this_roomName = r[23].value
            this_roomName = r[23].value
            this_type = 0

            if len(whichRooms) == 0 or this_roomID in whichRooms:
                print('Reading door #', this_ID)

                # Handle Forcing
                this_forced = r[41].value  # Column AP
                if this_forced is not None:
                    forcing[this_ID] = [this_forced]

                # Allow the option to declare which rooms to include
                if r[40].value is not None:
                    # This is an event (one-way) door that leads to a one-way exit
                    this_type = 1
                    addExit = True
                    if this_forced in doors:
                        # Check if it's connected to an entrance (forced destination) or an exit (force shared destination)
                        forced_type = [i for i in range(3) if this_forced in room_doors[door_rooms[this_forced]][i]][0]
                        if forced_type == 1:
                            # do not add a unique exit
                            addExit = False

                doors.append(this_ID)
                door_rooms[this_ID] = this_roomID
                door_descr[this_ID] = this_name
                if this_roomID not in rooms:
                    rooms.append(this_roomID)
                    # initialize room info
                    room_doors[this_roomID] = [[], [], []]
                    room_counts[this_roomID] = [0, 0, 0]
                room_doors[this_roomID][this_type].append(this_ID)
                room_counts[this_roomID][this_type] += 1

                if this_type == 1 and addExit:
                    # Add the one-way exit door
                    entr_ID = 1000 + this_ID  # create unique ID for associated 1-way entrance
                    entr_roomID = r[40].value  # NOTE: NEED TO ADD THIS FIELD (Col AO)!!!
                    doors.append(entr_ID)
                    door_rooms[entr_ID] = entr_roomID
                    door_descr[entr_ID] = this_name + ' DESTINATION'
                    if entr_roomID not in rooms:
                        rooms.append(entr_roomID)
                        # initialize room info
                        room_doors[entr_roomID] = [[], [], []]
                        room_counts[entr_roomID] = [0, 0, 0]
                    room_doors[entr_roomID][2].append(entr_ID)
                    room_counts[entr_roomID][2] += 1

    return room_doors, room_counts, door_descr, forcing


# zoneMaker: code to read in room lists and generate zones, one or more sets of fully-connected two-way doors that may
# only be entered or exited through one-ways.
def zoneMaker(room_doors, room_counts, forcing={}):
    # Generate list of valid (i.e. 2-way) doors & reverse door-->room lookup
    doors = []
    door_rooms = {}
    zones = []
    zone_counts = []
    for R in room_doors:
        doors.extend(room_doors[R][0])
        for d in room_doors[R][0]:
            door_rooms[d] = R
        # Each zone is a list of rooms in that zone.  Initially, each zone contains only one room
        zones.append([R])
        zone_counts.append(room_counts[R])
    door_zones = {}
    for zi in range(len(zones)):
        for d in room_doors[zones[zi][0]][0]:
            door_zones[d] = zi

    to_force = [d for d in forcing.keys() if d in doors]
    map = []
    # Connect all valid doors, creating zones in the process
    while len(doors) > 0:
        print('\n\nZone state:')
        for z in range(len(zones)):
            print(zones[z], ': ', zone_counts[z])
        print('Remaining doors: ', doors)

        # Get a list of dead end zones
        deadEnds = [zi for zi in range(len(zones)) if zone_counts[zi] == [1, 0, 0]]  # Identify the dead end zones
        print('Dead End zones: ', deadEnds)
        if len(to_force) > 0:
            door1 = to_force.pop(0)
            valid = forcing[door1]

        elif len(deadEnds) > 0:
            # First, always connect any dead-end zones (those with only one door)
            # Select a door in a dead end zone
            print('Zones of remaining doors:')
            for d in doors:
                print(d, door_zones[d])
            deadEndDoors = [d for d in doors if door_zones[d] in deadEnds]

            # Choose a random door
            door1 = deadEndDoors.pop(randrange(len(deadEndDoors)))
            doors.remove(door1)  # clean up
            zone1 = [i for i in range(len(zones)) if door_rooms[door1] in zones[i]][0]

            # Construct a list of valid zone connections:  Any zone that is not [1, 0, 0]; [1, 1, 0]; [1, 0, 1]
            valid_zone2 = [zi for zi in range(len(zones)) if zone_counts[zi] != [1, 0, 0] and zone_counts[zi] != [1, 1, 0] and zone_counts[zi] != [1, 0, 1]]
            valid = [d for d in doors if door_zones[d] in valid_zone2]

        else:
            # All dead-end zones have been connected.
            # Connect all remaining doors, following the rule that (unless only 2 doors are left) each zone must have
            # at least 1 entry and 1 exit.

            # Choose a random door
            door1 = doors.pop(randrange(len(doors)))
            zone1 = [i for i in range(len(zones)) if door_rooms[door1] in zones[i]][0]

            # Construct list of valid doors: start with all doors, then remove invalid ones
            valid = [d for d in doors]
            invalid = []
            if len(doors) > 2:
                # Remove doors that would create a zone with zero exits or zero entrances
                # i.e. a zone with [0, n, 0], or [0, 0, n].
                z1_exits = zone_counts[zone1][0] + zone_counts[zone1][1]
                z1_enter = zone_counts[zone1][0] + zone_counts[zone1][2]

                for d in valid:
                    print('Checking: ', d, ' (', door_zones[d],')')
                    z2 = door_zones[d]
                    if zone1 == z2:
                        # Self connection will remove two entrances and two exits from the zone
                        if (z1_exits-2 < 1) or (z1_enter-2 < 1):
                            # Creates a zone with no exit or no entrance
                            print('\tInvalid self: ', (z1_exits-2 < 1), (z1_enter-2 < 1))
                            invalid.append(d)
                    else:
                        z2_exits = zone_counts[z2][0] + zone_counts[z2][1]
                        z2_enter = zone_counts[z2][0] + zone_counts[z2][2]
                        # Note that the connection will remove 1 door (=1 exit + 1 entrance) from each zone
                        if ((z1_exits + z2_exits)-2 < 1) or ((z1_enter + z2_enter)-2 < 1):
                            # Creates a zone with no exit or no entrance
                            print('\tInvalid: ', ((z1_exits + z2_exits)-2 < 1), ((z1_enter + z2_enter)-2 < 1))
                            invalid.append(d)

            for i in invalid:
                valid.remove(i)

        print('Valid options for ', door1, '(', door_rooms[door1], '):\n', valid)

        # Select a connecting door
        door2 = valid.pop(randrange(len(valid)))
        zone2 = [i for i in range(len(zones)) if door_rooms[door2] in zones[i]][0]
        doors.remove(door2)

        # Write the connection
        map.append([door1, door2])
        print('Connected: ', door1, ' (zone ',zone1,') --> ', door2, '(zone ', zone2,')')

        # Modify the zones if necessary
        if zone1 != zone2:
            # Add zone2 to zone1
            zones[zone2].extend(zones[zone1])
            zones[zone1] = []

            # Adjust counts
            for i in range(len(zone_counts[zone2])):
                zone_counts[zone2][i] += zone_counts[zone1][i]
                zone_counts[zone1][i] = 0

            # Update door_zone listing
            for d in room_doors[door_rooms[door1]][0]:
                door_zones[d] = zone2
            #door_zones[door2] = zone2  # shouldn't be necessary

        # Decrement these two doors from the zone
        zone_counts[zone2][0] += -2

    # Clean up
    keep = [i for i in range(len(zones)) if zones[i]!=[]]
    zones = [zones[k] for k in keep]
    zone_counts = [zone_counts[k] for k in keep]

    return map, zones, zone_counts


# zoneWalker: code to read in a list of zones (one or more sets of fully-connected two-way doors that may
# only be entered or exited through one-ways) and connect their one-ways to make a single, fully-traversable final zone
def zoneWalker(room_doors, zones, zone_counts, forcing={}):
    # Generate lists of 1-way gates & reverse gate-->zone lookups
    nobs = []  # "outs" one-way exits
    nibs = []  # "ins" one-way entrances
    nob_rooms = {}
    nib_rooms = {}
    for R in room_doors:
        nobs.extend(room_doors[R][1])
        for nob in room_doors[R][1]:
            nob_rooms[nob] = R
        nibs.extend(room_doors[R][2])
        for nib in room_doors[R][2]:
            nib_rooms[nib] = R

    nob_zones = {}
    nib_zones = {}
    zone_nobs = {}
    zone_nibs = {}
    for zi in range(len(zones)):
        zone_nobs[zi] = []
        zone_nibs[zi] = []
        for R in zones[zi]:
            for nob in room_doors[R][1]:
                nob_zones[nob] = zi
                zone_nobs[zi].append(nob)
            for nib in room_doors[R][2]:
                nib_zones[nib] = zi
                zone_nibs[zi].append(nib)

    to_force = [n for n in forcing.keys() if n in nobs]
    forced_nibs = []
    for n in to_force:
        # Disallow forced nibs, except for the forcing nob
        if forcing[n] in nibs:
            forced_nibs.extend(forcing[n])

    # Walk through all valid one-ways, connecting all zones and returning to the starting point
    map = []
    walk = []
    if len(nobs) > 0:
        if len(to_force) > 0:
            # Start with a forced nob (reduces errors)
            nob = to_force[0]
            nobs.remove(nob)
        else:
            # Choose a random trap door to begin
            nob = nobs.pop(randrange(len(nobs)))

        zone1 = nob_zones[nob]
        walk.append(zone1) # record the path of the walk

        while len(nibs) > 0:
            print('\n\nWalk state:', walk)
            print('\nZone state:')
            for z in range(len(zones)):
                print(zones[z], ': ', zone_counts[z])
            print('Now Connecting: ', nob, '(', nob_zones[nob], ')')
            print('Remaining entrances: ', nibs)

            # Construct list of valid entrances: start with all nibs, then remove invalid ones
            valid = [n for n in nibs]
            invalid = []

            is_forced = nob in forcing.keys()
            if is_forced:
                # This is a forced connection.  There are two uses of this:
                # 1. forcing[nob] = [forced nibs]   # Only the entrances (nibs) in forcing[nob] are valid.
                # 2. forcing[nob] = [forced nobs]   # Force one exit to equal another exit (e.g. Umaro's cave #2 trapdoors)
                # We check whether forcing[nob] is entrances (nib) or exits (nob);
                #   If it's a list of entrances, connect one of them;
                #   if it's a list of exits, look to see if one of them is already connected;
                #       if so, connect this one to that same entrance (don't decrement entrances in zone 2)
                #       if not, just connect this one as normal.
                print('Found a forced connection...')
                forced = forcing[nob]
                these_nibs = [n for n in forced if n in nibs]
                if len(these_nibs) > 0:
                    # this is a list of entrances that are valid for this exit
                    print('\tSelecting only forced nibs:', these_nibs)
                    valid = these_nibs
                else:
                    # forced is a list of exits that must match.
                    # Look to see if any have been assigned.
                    print('\tLooking for previously-assigned paired exits... ')
                    assigned = [n for n in forced if n not in nobs]
                    if len(assigned) > 0:
                        # Find the previously-assigned connection in the map; take the nib
                        mapped_nib = [m for m in map if assigned[0] == m[0]][0][1]
                        print('\tFound one: ', assigned, ' --> ', mapped_nib)
                        valid = [mapped_nib]
                    else:
                        # Assign this one normally
                        print('\tNone found.  Proceeding normally.')
                        is_forced = False

            if len(nibs) > 1 and not is_forced:
                # Remove doors that would create a zone with zero exits or zero entrances
                # i.e. a zone with [0, n, 0], or [0, 0, n].
                z1_exits = zone_counts[zone1][1]
                z1_enter = zone_counts[zone1][2]
                invalid.extend(forced_nibs)  # don't allow connections to forced entrances

                for n in valid:
                    print('Checking ', n, '... ')
                    z2 = nib_zones[n]

                    # We need to remove loops into the walk that have no downstream exits
                    if zone1 == z2:
                        # Connection will remove an exit and an entrance from the zone
                        print('\tis same zone:', zone1, '--> [', z1_exits, z1_enter,']')
                        if ((z1_exits - 1) < 1):
                            # Connection would create a walk with no exits # or no entrances
                            print('\tRemoving ', n)
                            invalid.append(n)

                    elif z2 in walk:
                        # Search for a remaining downstream exit
                        z2i = walk.index(z2)
                        flag = 0
                        for wi in range(z2i, len(walk)):
                            if walk[wi] == zone1 and zone_counts[zone1][1] > 1:
                                flag = 1
                                break
                            elif zone_counts[walk[wi]][1] > 0:
                                flag = 1
                                break
                        if flag == 0:
                            # If no remaining downstream exit, remove this entrance
                            invalid.append(n)

                    else:
                        # z2 isn't in the walk yet.  Verify that the connection would still have exits.
                        z2_exits = zone_counts[z2][1]
                        z2_enter = zone_counts[z2][2]
                        print('\tNot the same zone:', zone1, '--> [', z1_exits, z1_enter,'], ', z2, '--> [', z2_exits, z2_enter,']')
                        # Connection will remove 1 exit from zone1 and 1 entrance from zone2
                        if (z2_exits < 1):
                            # Connection would create a walk with no exits
                            print('\tRemoving ', n)
                            invalid.append(n)

            for i in invalid:
                valid.remove(i)

            # Select an entrance to connect:
            print('Valid entrances: ', valid)
            nib = valid.pop(randrange(len(valid)))
            zone2 = nib_zones[nib]
            if nib in nibs:
                nibs.remove(nib)
                zone_counts[zone2][2] -= 1  # decrement entrances (nibs) in zone2
            if nib in forced_nibs:
                forced_nibs.remove(nib)

            # Write the connection
            map.append([nob, nib])
            print('Connected: ', nob, ' (zone ', zone1, ') --> ', nib, '(zone ', zone2, ')')
            zone_counts[zone1][1] -= 1  # decrement exits (nobs) in zone1

            # Update the walk and zone counts
            walk.append(zone2)  # DOES it matter than a zone can appear more than once in the walk?
            # No, this is good: say the walk is ['A', 'B', 'C', 'D', 'A', 'B'].
            # The downstream search always starts from the first instance (index = 1) so it will repeat any repeated
            # rooms (index = 5), but this just confirms you are always checking every zone you can reach for exits
            # (including 'A' @ index = 4; although A is technically upstream of B, B is also upstream of A).

            # If we created a loop, combine all zones in the loop into a new zone
            # A loop is created when a zone appears a second time in the walk.  It will always be bookended by zone2.
            if walk.count(zone2) > 1:
                loop = walk[walk.index(zone2):-1]
                print('Found a loop: ', loop)
                if len(loop) == 1:
                    print('Self-only loop, skipping.')
                    walk = walk[:-1]
                else:
                    # Create a new zone with the properties of all the zones in the loop
                    print('Compressing loop.')
                    # Gather information on the new zone
                    newzone = []
                    newzone_count = [0, 0, 0]
                    newzone_nobs = []
                    newzone_nibs = []
                    for zi in loop:
                        newzone.extend(zones[zi])
                        for j in range(3):
                            newzone_count[j] += zone_counts[zi][j]
                        newzone_nobs.extend(zone_nobs[zi])
                        newzone_nibs.extend(zone_nibs[zi])
                        # Delete old zone information
                        zones[zi] = []
                        zone_counts[zi] = [0,0,0]
                        zone_nobs[zi] = []
                        zone_nibs[zi] = []
                    nzi = len(zones)  # new zone index
                    # Create the new zone
                    zones.append(newzone)
                    zone_counts.append(newzone_count)
                    zone_nobs[nzi] = newzone_nobs
                    # Update dictionaries for nob zones & nib zones
                    for n in zone_nobs[nzi]:
                        nob_zones[n] = nzi
                    zone_nibs[nzi] = newzone_nibs
                    for n in zone_nibs[nzi]:
                        nib_zones[n] = nzi
                    # Update walk to replace the loop with the new zone ID
                    walk = walk[:walk.index(zone2)]
                    walk.append(nzi)
                    zone2 = nzi

            # Construct the list of all downstream exits...
            z2i = walk.index(zone2)
            available = []
            for wi in range(z2i, len(walk)):
                available.extend([nob for nob in zone_nobs[walk[wi]] if nob in nobs])
            # ... And randomly select one to connect:
            available = list(set(available))  # just unique values
            print('Available exits: ', available)
            if len(available) > 0:
                nob = available.pop(randrange(len(available)))
                nobs.remove(nob)
                zone1 = nob_zones[nob]
            else:
                if len(nobs) > 0:
                    print('ERROR: invalid situation encountered!')
                    break

    return map, walk
